from __future__ import annotations

from collections import defaultdict
from dataclasses import asdict, dataclass
from pathlib import Path

import gdstk
from openpyxl import Workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string, get_column_letter

from .base import BaseEngine


@dataclass
class FilmVolumeRow:
    layer: int
    datatype: int
    total_area: float


class FilmVolumeEngine(BaseEngine):
    HEADERS = ("Layer", "Datatype", "Total Area")
    MERGE_PRECISION = 1e-3
    AREA_DECIMALS = 6

    def __init__(self):
        super().__init__()
        self._loaded_gds_path: str | None = None

    def load_lib(self, path) -> list[str]:
        cell_names = super().load_lib(path)
        self._loaded_gds_path = str(Path(path).resolve())
        return cell_names

    def load_target_cell(self, gds_path, cell_name: str) -> gdstk.Cell:
        resolved_path = str(Path(gds_path).resolve())
        if self._loaded_gds_path != resolved_path or cell_name not in self.cells_map:
            self.load_lib(gds_path)

        if cell_name not in self.cells_map:
            raise ValueError(f"Cell '{cell_name}' 未找到")

        return self.cells_map[cell_name]

    def load_flat_cell(self, gds_path, cell_name: str) -> gdstk.Cell:
        return self.load_target_cell(gds_path, cell_name).flatten()

    def summarize_cell_areas(
        self,
        gds_path,
        cell_name: str,
        included_layers: set[int] | None = None,
    ) -> list[FilmVolumeRow]:
        target_cell = self.load_target_cell(gds_path, cell_name)
        rows = self.summarize_cell_geometry_areas(target_cell, included_layers=included_layers)
        return rows

    def summarize_flat_cell_areas(
        self,
        flat_cell: gdstk.Cell,
        included_layers: set[int] | None = None,
    ) -> list[FilmVolumeRow]:
        return self.summarize_cell_geometry_areas(flat_cell, included_layers=included_layers)

    def summarize_cell_geometry_areas(
        self,
        cell: gdstk.Cell,
        included_layers: set[int] | None = None,
    ) -> list[FilmVolumeRow]:
        rows = self._summarize_merged_areas(cell, included_layers=included_layers)
        if not rows:
            raise ValueError("目标 Cell 及其子层级里没有可统计的图形面积")
        return rows

    def write_area_table(
        self,
        rows: list[FilmVolumeRow],
        output_path,
        sheet_name: str = "AreaSummary",
        start_cell: str = "A1",
    ) -> dict:
        if not rows:
            raise ValueError("没有可写入 Excel 的面积数据")

        normalized_sheet_name = self._normalize_sheet_name(sheet_name)
        start_row, start_column, normalized_start_cell = self._parse_start_cell(start_cell)

        workbook = Workbook()
        try:
            worksheet = self._prepare_worksheet(workbook, normalized_sheet_name)

            current_row = start_row
            for offset, header in enumerate(self.HEADERS):
                worksheet.cell(current_row, start_column + offset, header)
            current_row += 1

            for item in rows:
                worksheet.cell(current_row, start_column, item.layer)
                worksheet.cell(current_row, start_column + 1, item.datatype)
                area_cell = worksheet.cell(current_row, start_column + 2, item.total_area)
                area_cell.number_format = "0.000000"
                current_row += 1

            self._adjust_column_widths(worksheet, start_column)
            workbook.save(output_path)
        finally:
            workbook.close()

        return {
            "output_path": str(output_path),
            "sheet_name": normalized_sheet_name,
            "start_cell": normalized_start_cell,
            "row_count": len(rows),
        }

    @staticmethod
    def rows_to_dicts(rows: list[FilmVolumeRow]) -> list[dict]:
        return [asdict(row) for row in rows]

    def _summarize_merged_areas(
        self,
        cell: gdstk.Cell,
        included_layers: set[int] | None = None,
    ) -> list[FilmVolumeRow]:
        polygons_by_spec: dict[tuple[int, int], list[gdstk.Polygon]] = defaultdict(list)
        for polygon in cell.get_polygons(apply_repetitions=True, include_paths=True):
            if included_layers is not None and polygon.layer not in included_layers:
                continue
            polygons_by_spec[(polygon.layer, polygon.datatype)].append(polygon)

        rows: list[FilmVolumeRow] = []
        for (layer, datatype), polygons in polygons_by_spec.items():
            merged_polygons = gdstk.boolean(
                polygons,
                [],
                "or",
                precision=self.MERGE_PRECISION,
                layer=layer,
                datatype=datatype,
            )
            total_area = sum(float(polygon.area()) for polygon in merged_polygons)
            if total_area <= 0:
                continue
            rows.append(
                FilmVolumeRow(
                    layer=layer,
                    datatype=datatype,
                    total_area=round(total_area, self.AREA_DECIMALS),
                )
            )

        rows.sort(key=lambda item: (item.layer, item.datatype))
        return rows

    @staticmethod
    def _normalize_sheet_name(sheet_name: str) -> str:
        cleaned = str(sheet_name or "").strip()
        for char in "\\/*?:[]":
            cleaned = cleaned.replace(char, "_")
        cleaned = cleaned[:31].strip()
        return cleaned or "AreaSummary"

    @staticmethod
    def _parse_start_cell(start_cell: str) -> tuple[int, int, str]:
        candidate = str(start_cell or "").strip().upper() or "A1"
        try:
            column_letters, row_number = coordinate_from_string(candidate)
            column_number = column_index_from_string(column_letters)
        except ValueError as exc:
            raise ValueError("起始单元格必须是类似 A1 的格式") from exc

        if row_number < 1 or column_number < 1:
            raise ValueError("起始单元格必须从第 1 行第 1 列开始")

        normalized = f"{get_column_letter(column_number)}{row_number}"
        return row_number, column_number, normalized

    @staticmethod
    def _prepare_worksheet(workbook, sheet_name: str):
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]

        active_sheet = workbook.active
        if (
            len(workbook.sheetnames) == 1
            and active_sheet.title == "Sheet"
            and active_sheet.max_row == 1
            and active_sheet.max_column == 1
            and active_sheet["A1"].value is None
        ):
            active_sheet.title = sheet_name
            return active_sheet

        return workbook.create_sheet(title=sheet_name)

    @staticmethod
    def _adjust_column_widths(worksheet, start_column: int) -> None:
        widths = {
            start_column: 12,
            start_column + 1: 12,
            start_column + 2: 18,
        }
        for column_index, width in widths.items():
            worksheet.column_dimensions[get_column_letter(column_index)].width = max(
                worksheet.column_dimensions[get_column_letter(column_index)].width or 0,
                width,
            )
