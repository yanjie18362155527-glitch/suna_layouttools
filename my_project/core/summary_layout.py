from __future__ import annotations

import re
import shutil
import tempfile
from dataclasses import asdict, dataclass
from pathlib import Path

import gdstk
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .base import BaseEngine


@dataclass
class SummaryLayoutEntry:
    column: str
    database: str
    top_cell: str
    array_columns: int
    array_rows: int
    center_x: float
    center_y: float


@dataclass
class ImportedCellBundle:
    source_path: Path
    source_database: str
    requested_top_cell_name: str
    resolved_top_cell_name: str
    imported_top: gdstk.Cell
    unit: float
    precision: float


class SummaryLayoutEngine(BaseEngine):

    def list_sheet_names(self, workbook_path) -> list[str]:
        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        try:
            return workbook.sheetnames
        finally:
            workbook.close()

    def parse_sheet(self, workbook_path, sheet_name: str) -> list[SummaryLayoutEntry]:
        # This workflow does a lot of random-access cell reads. In openpyxl,
        # read_only mode is optimized for streaming rows, but becomes much
        # slower for repeated worksheet.cell(...) lookups.
        workbook = load_workbook(workbook_path, data_only=True, read_only=False)
        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' does not exist")

            worksheet = workbook[sheet_name]
            database_anchor = self._find_label(worksheet, "Database")
            if database_anchor is None:
                raise ValueError("Database row not found")

            top_anchor = self._find_nearby_label(worksheet, "Top cell", database_anchor[0], database_anchor[1])
            chip_array_anchor = self._find_chip_array_anchor(worksheet, database_anchor[0], database_anchor[1])
            center_anchor = self._find_center_location_anchor(worksheet, database_anchor[0], database_anchor[1])
            if top_anchor is None or chip_array_anchor is None or center_anchor is None:
                raise ValueError(
                    "Top cell, Chip array, or 4X / 5X LBC center location block was not found"
                )

            array_y_row = chip_array_anchor[0] + 1
            center_y_row = center_anchor[0] + 1
            if array_y_row > worksheet.max_row or center_y_row > worksheet.max_row:
                raise ValueError("The selected sheet is missing the y row for Chip array or center location")

            entries: list[SummaryLayoutEntry] = []
            for column_index in range(database_anchor[1] + 1, worksheet.max_column + 1):
                database_name = self._normalize_text(worksheet.cell(database_anchor[0], column_index).value)
                top_cell_name = self._normalize_text(worksheet.cell(top_anchor[0], column_index).value)
                array_columns = self._to_int(worksheet.cell(chip_array_anchor[0], column_index).value, default=1)
                array_rows = self._to_int(worksheet.cell(array_y_row, column_index).value, default=1)
                center_x = self._to_float(worksheet.cell(center_anchor[0], column_index).value)
                center_y = self._to_float(worksheet.cell(center_y_row, column_index).value)

                if not database_name:
                    continue
                if self._normalize_label(database_name) == "database":
                    continue
                if not top_cell_name:
                    continue
                if center_x is None or center_y is None:
                    continue

                entries.append(
                    SummaryLayoutEntry(
                        column=self._column_letter(column_index),
                        database=database_name,
                        top_cell=top_cell_name,
                        array_columns=max(array_columns, 1),
                        array_rows=max(array_rows, 1),
                        center_x=center_x,
                        center_y=center_y,
                    )
                )

            if not entries:
                raise ValueError("No placement entries were parsed from the selected sheet")

            return entries
        finally:
            workbook.close()

    def process(
        self,
        workbook_path,
        sheet_name: str,
        gds_paths: list[Path],
        output_path,
        output_top_name: str,
        street_width: float = 0.0,
        entries: list[SummaryLayoutEntry] | None = None,
    ) -> dict:
        if entries is None:
            entries = self.parse_sheet(workbook_path, sheet_name)
        if not gds_paths:
            raise ValueError("No GDS files were uploaded")

        normalized_paths = [Path(path).expanduser() for path in gds_paths]
        valid_paths = [path for path in normalized_paths if path.exists() and path.is_file() and path.suffix.lower() == ".gds"]
        file_index = self._build_uploaded_gds_index(valid_paths)
        if not file_index:
            raise ValueError("No valid .gds files were uploaded")

        layout_top_name = self.sanitize_cell_name(output_top_name or sheet_name or "summary_layout")
        self.lib = gdstk.Library()
        self.cells_map = {}
        layout_cell = self.lib.new_cell(layout_top_name)
        self.cells_map[layout_top_name] = layout_cell

        imported_cache: dict[tuple[str, str], ImportedCellBundle] = {}
        duplicate_name_warnings = self._build_duplicate_file_warnings(file_index)
        unit_mismatch_files: set[str] = set()
        first_unit: float | None = None
        first_precision: float | None = None

        results = []
        placed_count = 0
        placed_instance_count = 0

        for entry in entries:
            resolved_path, match_note = self._resolve_database_path(entry.database, file_index)
            if resolved_path is None:
                results.append(
                        {
                            **asdict(entry),
                            "status": "skipped_missing_file",
                            "message": "Matching GDS was not found in uploaded files",
                            "matched_file": "",
                        }
                    )
                continue

            cache_key = (str(resolved_path.resolve()).lower(), entry.top_cell)
            if cache_key not in imported_cache:
                bundle = self._import_top_cell_bundle(resolved_path, entry.database, entry.top_cell, len(imported_cache) + 1)
                if bundle is None:
                    results.append(
                        {
                            **asdict(entry),
                            "status": "skipped_missing_top",
                            "message": f"Top cell '{entry.top_cell}' was not found in {resolved_path.name}",
                            "matched_file": resolved_path.name,
                        }
                    )
                    continue

                imported_cache[cache_key] = bundle

                if first_unit is None or first_precision is None:
                    first_unit = bundle.unit
                    first_precision = bundle.precision
                    self.lib.unit = bundle.unit
                    self.lib.precision = bundle.precision
                elif abs(bundle.unit - first_unit) > 1e-18 or abs(bundle.precision - first_precision) > 1e-18:
                    unit_mismatch_files.add(resolved_path.name)

            bundle = imported_cache[cache_key]
            bbox = bundle.imported_top.bounding_box()
            if bbox is None:
                results.append(
                    {
                        **asdict(entry),
                        "status": "skipped_empty_top",
                        "message": f"Top cell '{entry.top_cell}' has no usable bounding box",
                        "matched_file": resolved_path.name,
                    }
                )
                continue

            cell_width = bbox[1][0] - bbox[0][0]
            cell_height = bbox[1][1] - bbox[0][1]
            bbox_center_x = (bbox[0][0] + bbox[1][0]) / 2
            bbox_center_y = (bbox[0][1] + bbox[1][1]) / 2
            street_pitch = max(street_width, 0.0)
            column_spacing = cell_width + street_pitch
            row_spacing = cell_height + street_pitch

            # Excel stores the center of the first cell in the array:
            # the left-most, bottom-most instance. We therefore align the
            # first reference origin to that single-cell center, instead of
            # centering the whole array around the coordinate.
            origin_x = entry.center_x - bbox_center_x
            origin_y = entry.center_y - bbox_center_y

            if entry.array_columns == 1 and entry.array_rows == 1:
                layout_cell.add(gdstk.Reference(bundle.imported_top, origin=(origin_x, origin_y)))
            else:
                layout_cell.add(
                    gdstk.Reference(
                        bundle.imported_top,
                        origin=(origin_x, origin_y),
                        columns=entry.array_columns,
                        rows=entry.array_rows,
                        spacing=(column_spacing, row_spacing),
                    )
                )

            placed_count += 1
            placed_instance_count += entry.array_columns * entry.array_rows

            message_parts = ["placed"]
            if entry.array_columns > 1 or entry.array_rows > 1:
                message_parts.append(f"array {entry.array_columns}x{entry.array_rows}")
            if match_note:
                message_parts.append(match_note)
            if bundle.resolved_top_cell_name != entry.top_cell:
                message_parts.append(f"used fallback top cell {bundle.resolved_top_cell_name}")

            results.append(
                {
                    **asdict(entry),
                    "status": "placed",
                    "message": "; ".join(message_parts),
                    "matched_file": resolved_path.name,
                }
            )

        if placed_count == 0:
            raise ValueError("No GDS entries were placed successfully")

        self.save_lib(output_path)
        return {
            "sheet_name": sheet_name,
            "output_top_name": layout_top_name,
            "output_path": str(output_path),
            "total_entries": len(entries),
            "placed_count": placed_count,
            "placed_instance_count": placed_instance_count,
            "skipped_missing_file": sum(1 for item in results if item["status"] == "skipped_missing_file"),
            "skipped_missing_top": sum(1 for item in results if item["status"] == "skipped_missing_top"),
            "skipped_empty_top": sum(1 for item in results if item["status"] == "skipped_empty_top"),
            "duplicate_name_warnings": duplicate_name_warnings,
            "unit_mismatch_files": sorted(unit_mismatch_files),
            "results": results,
        }

    @staticmethod
    def sanitize_cell_name(raw_name: str) -> str:
        text = SummaryLayoutEngine._normalize_text(raw_name) or "summary_layout"
        text = re.sub(r"[^\w]+", "_", text)
        text = re.sub(r"_+", "_", text).strip("_")
        return text or "summary_layout"

    @staticmethod
    def entries_to_rows(entries: list[SummaryLayoutEntry]) -> list[dict]:
        return [asdict(entry) for entry in entries]

    @staticmethod
    def _normalize_text(value) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _normalize_label(value) -> str:
        text = SummaryLayoutEngine._normalize_text(value).lower().replace("\n", " ")
        return re.sub(r"\s+", " ", text).strip()

    @staticmethod
    def _to_float(value) -> float | None:
        if value is None or value == "":
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _to_int(value, default: int | None = None) -> int | None:
        if value is None or value == "":
            return default
        try:
            return int(round(float(value)))
        except (TypeError, ValueError):
            return default

    @staticmethod
    def _column_letter(index: int) -> str:
        letters = []
        value = index
        while value > 0:
            value, remainder = divmod(value - 1, 26)
            letters.append(chr(65 + remainder))
        return "".join(reversed(letters))

    @staticmethod
    def _find_label(worksheet: Worksheet, label: str) -> tuple[int, int] | None:
        normalized_target = SummaryLayoutEngine._normalize_label(label)
        for row in worksheet.iter_rows():
            for cell in row:
                if SummaryLayoutEngine._normalize_label(cell.value) == normalized_target:
                    return cell.row, cell.column
        return None

    @staticmethod
    def _find_nearby_label(
        worksheet: Worksheet,
        label: str,
        anchor_row: int,
        anchor_column: int,
        row_window: int = 4,
        column_window: int = 4,
    ) -> tuple[int, int] | None:
        normalized_target = SummaryLayoutEngine._normalize_label(label)
        best_match = None
        best_score = None

        min_row = max(1, anchor_row - 1)
        max_row = min(worksheet.max_row, anchor_row + row_window)
        min_column = max(1, anchor_column - 1)
        max_column = min(worksheet.max_column, anchor_column + column_window)

        for row_index in range(min_row, max_row + 1):
            for column_index in range(min_column, max_column + 1):
                if SummaryLayoutEngine._normalize_label(worksheet.cell(row_index, column_index).value) != normalized_target:
                    continue

                score = abs(row_index - anchor_row) * 100 + abs(column_index - anchor_column)
                if best_score is None or score < best_score:
                    best_match = (row_index, column_index)
                    best_score = score

        return best_match

    @staticmethod
    def _find_chip_array_anchor(worksheet: Worksheet, anchor_row: int, anchor_column: int) -> tuple[int, int] | None:
        return SummaryLayoutEngine._find_block_anchor(
            worksheet,
            anchor_row,
            anchor_column,
            predicate=lambda normalized: "chip array" in normalized,
        )

    @staticmethod
    def _find_center_location_anchor(worksheet: Worksheet, anchor_row: int, anchor_column: int) -> tuple[int, int] | None:
        primary = SummaryLayoutEngine._find_block_anchor(
            worksheet,
            anchor_row,
            anchor_column,
            predicate=lambda normalized: (
                "center location" in normalized
                and "2x" not in normalized
                and ("4x" in normalized or "5x" in normalized)
            ),
        )
        if primary is not None:
            return primary

        return SummaryLayoutEngine._find_block_anchor(
            worksheet,
            anchor_row,
            anchor_column,
            predicate=lambda normalized: "center location" in normalized and "2x" not in normalized,
        )

    @staticmethod
    def _find_block_anchor(
        worksheet: Worksheet,
        anchor_row: int,
        anchor_column: int,
        predicate,
        row_window: int = 24,
        column_window: int = 2,
    ) -> tuple[int, int] | None:
        best_match = None
        best_score = None

        min_row = max(1, anchor_row)
        max_row = min(worksheet.max_row, anchor_row + row_window)
        min_column = max(1, anchor_column - column_window)
        max_column = min(worksheet.max_column, anchor_column + column_window)

        for row_index in range(min_row, max_row + 1):
            for column_index in range(min_column, max_column + 1):
                raw_value = worksheet.cell(row_index, column_index).value
                if not isinstance(raw_value, str):
                    continue

                normalized = SummaryLayoutEngine._normalize_label(raw_value)
                if not predicate(normalized):
                    continue

                score = abs(row_index - anchor_row) * 100 + abs(column_index - anchor_column)
                if best_score is None or score < best_score:
                    best_match = (row_index, column_index)
                    best_score = score

        return best_match

    @staticmethod
    def _build_uploaded_gds_index(gds_paths: list[Path]) -> dict[str, list[Path]]:
        index: dict[str, list[Path]] = {}
        for path in gds_paths:
            index.setdefault(path.name.lower(), []).append(path)
        return index

    @staticmethod
    def _build_duplicate_file_warnings(file_index: dict[str, list[Path]]) -> list[str]:
        warnings = []
        for name, paths in sorted(file_index.items()):
            if len(paths) > 1:
                warnings.append(f"{name}: matched multiple files, using {paths[0]}")
        return warnings

    @staticmethod
    def _resolve_database_path(database_name: str, file_index: dict[str, list[Path]]) -> tuple[Path | None, str]:
        raw_name = Path(database_name).name
        stem = Path(raw_name).stem
        suffix = Path(raw_name).suffix.lower()

        candidates = [raw_name]
        if suffix != ".gds":
            candidates.append(f"{stem}.gds")

        checked = set()
        for candidate in candidates:
            normalized = candidate.lower()
            if normalized in checked:
                continue
            checked.add(normalized)
            if normalized in file_index:
                note = ""
                if candidate.lower() != raw_name.lower():
                    note = f"matched by fallback name {candidate}"
                return file_index[normalized][0], note

        return None, ""

    def _import_top_cell_bundle(
        self,
        gds_path: Path,
        database_name: str,
        top_cell_name: str,
        import_index: int,
    ) -> ImportedCellBundle | None:
        source_lib = self._read_library_with_safe_path(gds_path)
        source_cells = {cell.name: cell for cell in source_lib.cells}

        if top_cell_name in source_cells:
            top_cell = source_cells[top_cell_name]
            resolved_top_cell_name = top_cell_name
        else:
            top_levels = source_lib.top_level()
            if len(top_levels) != 1:
                return None
            top_cell = top_levels[0]
            resolved_top_cell_name = top_cell.name

        relevant_cells = {top_cell.name: top_cell}
        for dependency in top_cell.dependencies(True):
            if isinstance(dependency, gdstk.Cell):
                relevant_cells.setdefault(dependency.name, dependency)

        prefix = f"IMP{import_index:03d}__"
        imported_cells: list[gdstk.Cell] = []
        for original_name, cell in list(relevant_cells.items()):
            source_lib.rename_cell(cell, f"{prefix}{original_name}")
            imported_cells.append(cell)

        self.lib.add(*imported_cells)
        imported_top = next(cell for cell in imported_cells if cell.name == f"{prefix}{resolved_top_cell_name}")
        self.cells_map[imported_top.name] = imported_top

        return ImportedCellBundle(
            source_path=gds_path,
            source_database=database_name,
            requested_top_cell_name=top_cell_name,
            resolved_top_cell_name=resolved_top_cell_name,
            imported_top=imported_top,
            unit=source_lib.unit,
            precision=source_lib.precision,
        )

    @staticmethod
    def _read_library_with_safe_path(gds_path: Path):
        resolved_path = Path(gds_path).resolve()
        path_text = str(resolved_path)
        if path_text.isascii():
            return gdstk.read_gds(path_text)

        with tempfile.TemporaryDirectory(prefix="layout_auto_tools_") as temp_dir:
            temp_path = Path(temp_dir) / resolved_path.name
            shutil.copy2(resolved_path, temp_path)
            return gdstk.read_gds(str(temp_path))
