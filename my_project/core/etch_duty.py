from __future__ import annotations

from dataclasses import asdict, dataclass

from openpyxl import Workbook

from .film_volume import FilmVolumeEngine


@dataclass
class EtchDutyRow:
    layer: int
    datatype: int
    total_area: float
    duty_cycle_percent: float


class EtchDutyEngine(FilmVolumeEngine):
    HEADERS = ("Layer", "Datatype", "Total Area", "Duty Cycle (%)")
    MERGE_PRECISION = 0.001

    def summarize_shot_duty(
        self,
        gds_path,
        cell_name: str,
        included_layers: set[int] | None = None,
    ) -> tuple[list[EtchDutyRow], float]:
        target_cell = self.load_target_cell(gds_path, cell_name)
        bbox = target_cell.bounding_box()
        if bbox is None:
            raise ValueError(f"Shot Cell '{cell_name}' 没有可统计的几何边界")

        cell_width = float(bbox[1][0] - bbox[0][0])
        cell_height = float(bbox[1][1] - bbox[0][1])
        cell_total_area = round(cell_width * cell_height, self.AREA_DECIMALS)
        if cell_total_area <= 0:
            raise ValueError(f"Shot Cell '{cell_name}' 的总面积无效")

        area_rows = self.summarize_cell_geometry_areas(target_cell, included_layers=included_layers)
        duty_rows = [
            EtchDutyRow(
                layer=row.layer,
                datatype=row.datatype,
                total_area=row.total_area,
                duty_cycle_percent=round(row.total_area / cell_total_area * 100, self.AREA_DECIMALS),
            )
            for row in area_rows
        ]
        return duty_rows, cell_total_area

    def write_duty_table(
        self,
        rows: list[EtchDutyRow],
        output_path,
        sheet_name: str = "DutyCycleSummary",
        start_cell: str = "A1",
    ) -> dict:
        if not rows:
            raise ValueError("没有可写入 Excel 的占空比数据")

        normalized_sheet_name = self._normalize_sheet_name(sheet_name)
        start_row, start_column, normalized_start_cell = self._parse_start_cell(start_cell)

        workbook = Workbook()
        try:
            worksheet = self._prepare_worksheet(workbook, normalized_sheet_name)

            current_row = start_row
            for offset, header in enumerate(self.HEADERS):
                worksheet.cell(current_row, start_column + offset, header)
            current_row += 1

            for item in rows:
                worksheet.cell(current_row, start_column, item.layer)
                worksheet.cell(current_row, start_column + 1, item.datatype)

                area_cell = worksheet.cell(current_row, start_column + 2, item.total_area)
                area_cell.number_format = "0.000000"

                duty_cell = worksheet.cell(current_row, start_column + 3, item.duty_cycle_percent)
                duty_cell.number_format = "0.000000"

                current_row += 1

            self._adjust_column_widths(worksheet, start_column)
            worksheet.column_dimensions[self._column_letter(start_column + 3)].width = 18
            workbook.save(output_path)
        finally:
            workbook.close()

        return {
            "output_path": str(output_path),
            "sheet_name": normalized_sheet_name,
            "start_cell": normalized_start_cell,
            "row_count": len(rows),
        }

    @staticmethod
    def rows_to_dicts(rows: list[EtchDutyRow]) -> list[dict]:
        return [asdict(row) for row in rows]

    @staticmethod
    def _column_letter(index: int) -> str:
        letters = []
        value = index
        while value > 0:
            value, remainder = divmod(value - 1, 26)
            letters.append(chr(65 + remainder))
        return "".join(reversed(letters))
